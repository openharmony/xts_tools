#!/usr/bin/env python3
# Copyright (c) 2026 Shenzhen Kaihong Digital Industry Development Co., Ltd.
"""Split 0803 UI-compare xlsx into snap / assert / manual markdown lists."""
from __future__ import annotations

import argparse
import re
from pathlib import Path

try:
    import openpyxl
except ImportError as e:
    raise SystemExit("need openpyxl") from e

ROOT = Path(__file__).resolve().parents[1]


def classify(num: str, name: str, desc: str) -> str:
    t = f"{num} {name} {desc}".upper()
    # manual / external deps
    if any(x in t for x in ["美团", "华为账号", "运动健康", "应用市场"]):
        return "manual"
    if any(x in t for x in ["NATIVE", "TDD", "HIDUMPER", "DUMP", "TRACE", "DFX", "内存泄漏", "MEMORY"]):
        return "assert"
    if any(x in t for x in ["UEC", "UIEXTENSION", "UIEXTENSIONCOMPONENT"]):
        return "assert"
    if any(x in t for x in ["PC_EVENT", "鼠标", "表冠", "屏幕朗读", "BACKWARDANALYSIS", "无障碍"]):
        return "assert"
    if any(x in t for x in ["OUTLIERS", "IDE报错"]):
        return "assert"
    if any(
        x in t
        for x in [
            "PROGRESS",
            "LOADING",
            "LOADPROGRESS",
            "PICKER",
            "BADGE",
            "CALENDAR",
            "DATEPICKER",
            "TEXTPICKER",
            "DARK",
            "COLOR_MODE",
            "深色",
            "PARALLELIZATION",
            "CONTENTMODIFIER",
            "ANIMATETO",
            "ENABLESMOOTHEFFECT",
            "RINGSTYLE",
        ]
    ):
        return "snap"
    # DC visual stress → snap if about 属性切换视觉
    if "DC_" in t and "压力" in name:
        return "snap"
    # default: assert for public-capability-heavy, else snap if 组件组
    if "公共能力" in desc:
        return "assert"
    if "组件" in desc:
        return "snap"
    return "manual"


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--xlsx",
        default="/root/aiSkill/develop/需求/0803-UI对比自动化用例实现.xlsx",
    )
    ap.add_argument("--out-dir", default=str(ROOT / "docs"))
    args = ap.parse_args()
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]
    buckets = {"snap": [], "assert": [], "manual": []}
    for r in range(2, ws.max_row + 1):
        name, num, desc, pre, steps, exp = [ws.cell(r, c).value for c in range(1, 7)]
        if not num:
            continue
        num_s = str(num).strip()
        name_s = str(name or "").strip()
        desc_s = str(desc or "").strip()
        bucket = classify(num_s, name_s, desc_s)
        buckets[bucket].append((num_s, name_s, desc_s, str(steps or "")[:200], str(exp or "")[:200]))

    def write_md(path: Path, title: str, rows: list) -> None:
        lines = [
            f"# {title}",
            "",
            f"合计 **{len(rows)}** 条（自 0803 xlsx 自动分流，可人工调整）。",
            "",
            "| 编号 | 名称 | 描述组 |",
            "|------|------|--------|",
        ]
        for num, name, desc, _steps, _exp in rows:
            name_esc = name.replace("|", "\\|")
            lines.append(f"| `{num}` | {name_esc} | {desc} |")
        lines.append("")
        path.write_text("\n".join(lines), encoding="utf-8")
        print(f"wrote {path} n={len(rows)}")

    write_md(out_dir / "0803_snap_cases.md", "0803 适合截图对比（Snap）", buckets["snap"])
    write_md(out_dir / "0803_assert_cases.md", "0803 非截图断言（Assert）", buckets["assert"])
    write_md(out_dir / "0803_manual_cases.md", "0803 暂缓/人工/外仓", buckets["manual"])
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
