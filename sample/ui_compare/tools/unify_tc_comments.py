#!/usr/bin/env python3
# Copyright (c) 2026 Shenzhen Kaihong Digital Industry Development Co., Ltd.
"""Unify @tc.* comment blocks above Hypium it() in ui_compare Suite files.

Canonical block (aligned with project need):
  @tc.number / @tc.name / @tc.desc / @tc.type / @tc.size / @tc.level

- @tc.desc: English description of what interface / API this case verifies
- Preserves CRLF if the file already uses CRLF
- Skips model/ and testability/pages/test orphans
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
CACHE = ROOT / "tools" / ".tc_case_titles.json"

IT_WITH_COMMENT = re.compile(
    r"(?P<indent>[ \t]*)/\*(?P<body>.*?)\*/\s*\n"
    r"(?P=indent)it\s*\(\s*(?P<q>['\"])(?P<itname>[^'\"]+)(?P=q)",
    re.DOTALL,
)

IT_BARE = re.compile(
    r"(?P<indent>[ \t]*)it\s*\(\s*(?P<q>['\"])(?P<itname>[^'\"]+)(?P=q)",
)

FIELD_RE = re.compile(
    r"@tc\.(number|name|type|size|level|desc)\s*:?\s*(.*)$",
    re.IGNORECASE | re.MULTILINE,
)

# Common token → English phrase for case-id based titles
TOKEN_MAP = {
    "COMPONENT": "component",
    "INFOMATION": "information",
    "INFORMATION": "information",
    "LOADPROGRESS": "LoadingProgress",
    "LOADINGPROGRESS": "LoadingProgress",
    "PROGRESS": "Progress",
    "ANIMATETO": "animateTo",
    "ENABLESMOOTHEFFECT": "enableSmoothEffect",
    "INTERFACE": "interface",
    "BUILDER": "contentModifier builder",
    "CONTENTMODIFIER": "contentModifier",
    "DARK": "dark",
    "COLOR": "color",
    "MODE": "mode",
    "MIRROR": "mirror layout",
    "ADVANCEDCOMPONENTS": "advanced components",
    "COMPONENTFUNCTION": "component function",
    "TEXTPICKER": "TextPicker",
    "DATEPICKER": "DatePicker",
    "CALENDARPICKER": "CalendarPicker",
    "TIMEPICKER": "TimePicker",
    "PICKER": "Picker",
    "BADGE": "Badge",
    "TEXTCLOCK": "TextClock",
    "OUTLIERS": "outlier parameters",
    "DISABLEDDATE": "disabledDateRange",
    "CUSTOMANIMATION": "custom animation",
    "PARALLELIZATION": "multi-thread parallelization",
    "NATIVE": "native memory tracing",
    "DISPLAY": "display",
    "PATTERNLOCK": "PatternLock",
    "UIEXTENSIONCOMPONENT": "UIExtensionComponent",
    "SPECIALCOMPONENTS": "special components",
    "MEDIA": "media",
    "BUTTON": "button",
    "IMAGE": "Image",
    "RING": "ring style",
    "SHADOW": "shadow",
    "SCAN": "scan effect",
    "EVENT": "event",
    "DFX": "DFX",
    "CONTAINER": "container",
    "LARGE": "large font / aging",
    "FONT": "font",
    "ARKUI": "ArkUI",
}


ZH_GLOSSARY = [
    (r"环形进度条", "ring Progress"),
    (r"进度条", "Progress"),
    (r"深色模式", "dark color mode"),
    (r"浅色", "light mode"),
    (r"适老化", "aging-friendly / large font"),
    (r"扫光", "scan effect"),
    (r"动效", "animation"),
    (r"异常值", "outlier / invalid values"),
    (r"空值", "null / empty values"),
    (r"Native内存追踪", "native memory tracing"),
    (r"内存追踪", "native memory tracing"),
    (r"内存泄漏", "memory leak"),
    (r"未设置宽高", "without width/height"),
    (r"PatternLock组件", "PatternLock "),
    (r"LoadingProgress组件", "LoadingProgress "),
    (r"Progress组件", "Progress "),
    (r"TextClock组件", "TextClock "),
    (r"CalendarPicker", "CalendarPicker"),
    (r"设置", "set "),
    (r"验证", "verify "),
    (r"测试", ""),
    (r"组件", " "),
    (r"接口", " interface"),
    (r"属性", " property"),
    (r"场景", " scenario"),
    (r"正常展示", "renders correctly"),
    (r"功能正常", "functions correctly"),
    (r"为true", " to true"),
    (r"为false", " to false"),
    (r"不设置", " unset"),
    (r"传入", " pass "),
]


def load_title_cache() -> dict[str, str]:
    if CACHE.is_file():
        try:
            return json.loads(CACHE.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def build_title_cache_from_xlsx(xlsx: Path) -> dict[str, str]:
    try:
        import openpyxl
    except ImportError:
        return {}
    if not xlsx.is_file():
        return {}
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out: dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        name, num = ws.cell(r, 1).value, ws.cell(r, 2).value
        if not num:
            continue
        out[str(num).strip()] = str(name or "").strip()
    CACHE.write_text(json.dumps(out, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return out


def zh_to_english_hint(zh: str) -> str:
    if not zh:
        return ""
    s = zh
    for pat, en in ZH_GLOSSARY:
        s = re.sub(pat, en, s)
    # drop leftover CJK blocks into compact placeholder
    if re.search(r"[\u4e00-\u9fff]", s):
        # keep mixed result but strip heavy CJK runs
        s = re.sub(r"[\u4e00-\u9fff]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip(" -_")
    return s


def tokens_from_case_id(case_id: str) -> list[str]:
    # strip common prefixes
    s = case_id
    for pref in (
        "SUB_ACE_UI_COMPONENT_",
        "SUB_ACE_UI_",
        "SUB_ACE_TOOLCHAIN_ARKUI_",
        "SUB_ACE_ACTION_ARKUI_",
        "SUB_ACE_",
        "ARKUI_",
    ):
        if s.startswith(pref):
            s = s[len(pref) :]
            break
    parts = [p for p in s.split("_") if p]
    # drop trailing pure digits like 0530 / 00031
    if parts and re.fullmatch(r"\d+", parts[-1]):
        parts = parts[:-1]
    return parts


def english_from_case_id(case_id: str) -> tuple[str, str]:
    parts = tokens_from_case_id(case_id)
    words: list[str] = []
    for p in parts:
        up = p.upper()
        if up in TOKEN_MAP:
            words.append(TOKEN_MAP[up])
        elif re.fullmatch(r"\d+", p):
            continue
        else:
            # Camel-ish: TEXTPICKER already handled; fallback lower
            words.append(p.lower())
    # dedupe consecutive
    cleaned: list[str] = []
    for w in words:
        if not cleaned or cleaned[-1] != w:
            cleaned.append(w)
    phrase = " ".join(cleaned).strip() or "UI component"
    name = phrase[0].upper() + phrase[1:] if phrase else case_id
    # shorten name
    if len(name) > 90:
        name = name[:87] + "..."
    desc = f"Verify {phrase} interface behavior and expected UI result for this case."
    return name, desc


def looks_like_case_id(s: str) -> bool:
    if not s:
        return True
    if s.startswith("SUB_") or s.startswith("ARKUI_"):
        return True
    if re.fullmatch(r"[A-Z0-9_]+", s) and "_" in s and len(s) > 20:
        return True
    return False


def looks_english(s: str) -> bool:
    if not s or looks_like_case_id(s):
        return False
    # mostly latin letters
    letters = re.findall(r"[A-Za-z]", s)
    cjk = re.findall(r"[\u4e00-\u9fff]", s)
    return len(letters) >= 3 and len(cjk) == 0


def normalize_block(
    itname: str,
    body: str,
    indent: str,
    title_map: dict[str, str],
) -> str:
    fields: dict[str, str] = {}
    for m in FIELD_RE.finditer(body):
        key = m.group(1).lower()
        val = m.group(2).strip()
        val = re.sub(r"\s*\*+\s*$", "", val).strip()
        if val:
            fields[key] = val

    number = fields.get("number") or itname
    zh = title_map.get(number) or title_map.get(itname) or ""
    id_name, id_desc = english_from_case_id(number)

    name = fields.get("name") or ""
    desc = fields.get("desc") or ""

    # name: prefer good English; else derive
    if looks_like_case_id(name) or not looks_english(name):
        if zh:
            hint = zh_to_english_hint(zh)
            name = hint[:90] if hint and looks_english(hint) else id_name
        else:
            name = id_name if looks_like_case_id(name) or not name else name
        if looks_like_case_id(name) or not name:
            name = id_name

    # desc: always English; prefer existing good desc; else build
    if not looks_english(desc):
        if zh:
            hint = zh_to_english_hint(zh)
            if hint and looks_english(hint):
                desc = f"Verify the interface under test: {hint}."
            else:
                desc = id_desc
                if zh:
                    # keep a stable machine desc even if glossary incomplete
                    desc = (
                        f"Verify the interface under test for scenario "
                        f"described by case '{number}'."
                    )
        else:
            desc = id_desc if not looks_english(desc) else desc
    if not desc:
        desc = id_desc

    typ = fields.get("type") or "Function"
    if typ.lower() in ("function test", "func", "function"):
        typ = "Function"
    size = fields.get("size") or "MediumTest"
    if size.lower() in ("uitest", "mediumtest", "medium"):
        size = "MediumTest"
    level = fields.get("level") or "3"
    level = re.sub(r"^LEVEL\s*", "", level, flags=re.I).strip() or "3"

    lines = [
        f"{indent}/*",
        f"{indent} * @tc.number : {number}",
        f"{indent} * @tc.name   : {name}",
        f"{indent} * @tc.desc   : {desc}",
        f"{indent} * @tc.type   : {typ}",
        f"{indent} * @tc.size   : {size}",
        f"{indent} * @tc.level  : {level}",
        f"{indent} */",
    ]
    return "\n".join(lines)


def process_text(text: str, title_map: dict[str, str]) -> tuple[str, int]:
    changed = 0

    def repl_comment(m: re.Match) -> str:
        nonlocal changed
        if "@tc." not in m.group("body"):
            return m.group(0)
        indent = m.group("indent")
        itname = m.group("itname")
        q = m.group("q")
        new_block = normalize_block(itname, m.group("body"), indent, title_map)
        old = m.group(0)
        new = f"{new_block}\n{indent}it({q}{itname}{q}"
        if new != old:
            changed += 1
        return new

    out = IT_WITH_COMMENT.sub(repl_comment, text)

    lines = out.splitlines(keepends=True)
    result: list[str] = []
    i = 0
    while i < len(lines):
        line = lines[i]
        m = IT_BARE.match(line.rstrip("\n\r"))
        if m:
            j = len(result) - 1
            while j >= 0 and result[j].strip() == "":
                j -= 1
            has_block = j >= 0 and "*/" in result[j]
            if not has_block:
                indent = m.group("indent")
                itname = m.group("itname")
                block = normalize_block(itname, "", indent, title_map) + "\n"
                result.append(block)
                changed += 1
        result.append(line)
        i += 1
    return "".join(result), changed


def write_preserving_newlines(path: Path, text: str, original: bytes) -> None:
    use_crlf = b"\r\n" in original
    data = text.encode("utf-8")
    data = data.replace(b"\r\n", b"\n").replace(b"\r", b"\n")
    if use_crlf:
        data = data.replace(b"\n", b"\r\n")
    path.write_bytes(data)


def iter_test_files(project: Path) -> list[Path]:
    test_roots = list(project.glob("**/entry/src/ohosTest/ets/test"))
    if not test_roots:
        return []
    files: list[Path] = []
    for test_root in test_roots:
        for p in test_root.rglob("*.test.ets"):
            norm = str(p).replace("\\", "/")
            if "/model/" in norm:
                continue
            if "/testability/pages/test/" in norm:
                continue
            files.append(p)
    return sorted(set(files))


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--project", action="append", help="project dir name or path")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--root", default=str(ROOT))
    ap.add_argument(
        "--xlsx",
        default="/root/aiSkill/develop/需求/0803-UI对比自动化用例实现.xlsx",
    )
    ap.add_argument("--rebuild-cache", action="store_true")
    args = ap.parse_args()
    root = Path(args.root)

    title_map = load_title_cache()
    if args.rebuild_cache or not title_map:
        title_map = build_title_cache_from_xlsx(Path(args.xlsx))
        # also merge docs tables
        for md in (root / "docs").glob("0803_*.md"):
            for line in md.read_text(encoding="utf-8", errors="ignore").splitlines():
                m = re.match(r"\|\s*`([^`]+)`\s*\|\s*([^|]+)\|", line)
                if m:
                    title_map.setdefault(m.group(1).strip(), m.group(2).strip())
        CACHE.write_text(
            json.dumps(title_map, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
        )

    if args.project:
        projects = []
        for p in args.project:
            path = Path(p)
            if not path.is_absolute():
                path = root / p
            projects.append(path)
    else:
        projects = sorted(root.glob("uiCompareTest_*")) + sorted(
            root.glob("uiAssertTest_*")
        )

    total_files = 0
    total_blocks = 0
    for proj in projects:
        if not proj.is_dir():
            print(f"skip missing {proj}", file=sys.stderr)
            continue
        for f in iter_test_files(proj):
            original = f.read_bytes()
            text = original.decode("utf-8", errors="replace")
            new, n = process_text(text, title_map)
            if n == 0 and "@tc.desc" in text:
                # still rewrite if desc missing in any block
                if re.search(r"@tc\.name\s*:", text) and not re.search(
                    r"@tc\.desc\s*:", text
                ):
                    new, n = process_text(text, title_map)
            # force rewrite when name==number style without desc
            if n == 0:
                # detect missing desc fields
                blocks = list(IT_WITH_COMMENT.finditer(text))
                need = False
                for m in blocks:
                    if "@tc.desc" not in m.group("body"):
                        need = True
                        break
                    # name equals itname
                    nm = re.search(r"@tc\.name\s*:?\s*(.+)", m.group("body"))
                    if nm and looks_like_case_id(nm.group(1).strip()):
                        need = True
                        break
                if need:
                    new, n = process_text(text, title_map)
                    if n == 0:
                        # force count by comparing normalized
                        forced, _ = process_text(text, title_map)
                        if forced != text:
                            new, n = forced, max(1, len(blocks))
            if new == text:
                continue
            # recount changed blocks roughly
            if n == 0:
                n = 1
            total_files += 1
            total_blocks += n
            print(f"{f.relative_to(root)}: updated")
            if not args.dry_run:
                write_preserving_newlines(f, new, original)
    print(f"done files={total_files} blocks~={total_blocks} dry_run={args.dry_run}")
    return 0


if __name__ == "__main__":
    # fix accidental walrus in FIELD_RE at import time for older py — already 3.10+
    raise SystemExit(main())
